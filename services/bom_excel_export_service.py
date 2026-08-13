from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class BomExcelExportService:
    """BOM 조회 결과를 메모리 기반 XLSX 파일로 생성합니다."""

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    INFO_FILL = PatternFill("solid", fgColor="D9EAF7")
    THIN_BORDER = Border(bottom=Side(style="thin", color="B4C6D7"))

    COLUMN_LABELS = {
        "level": "BOM Level",
        "bom_level": "BOM Level",
        "product_id": "제품 ID",
        "bom_parent": "Parent 자재",
        "bom_parent_name": "Parent 자재명",
        "bom_child": "자재 ID",
        "bom_child_name": "자재명",
        "material_id": "자재 ID",
        "material_name": "자재명",
        "material_type": "자재 유형",
        "quantity": "수량",
        "unit": "단위",
        "location": "Location",
        "sequence_no": "Sequence",
        "supplier_id": "공급사 ID",
        "supplier_name": "공급사",
        "unit_price": "단가",
        "currency": "통화",
        "effective_from": "유효 시작일",
        "effective_to": "유효 종료일",
        "start_date": "유효 시작일",
        "end_date": "유효 종료일",
        "root_model": "Root 모델",
        "bom_path": "BOM 경로",
        "required_quantity": "누적 소요량",
        "status": "상태",
    }

    @classmethod
    def build(
        cls,
        rows: list[dict],
        product_id: str,
        as_of_date: str | None,
        generated_at: str,
    ) -> bytes:
        if not rows:
            raise ValueError("Excel로 내보낼 BOM 조회 결과가 없습니다.")

        columns = list(rows[0].keys())
        workbook = Workbook()
        result_sheet = workbook.active
        result_sheet.title = "BOM 조회결과"
        result_sheet.freeze_panes = "A2"
        result_sheet.sheet_view.showGridLines = False

        for column_index, key in enumerate(columns, 1):
            cell = result_sheet.cell(1, column_index, cls.COLUMN_LABELS.get(key, key))
            cell.fill = cls.HEADER_FILL
            cell.font = Font(name="맑은 고딕", color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        result_sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

        for row_index, row in enumerate(rows, 2):
            for column_index, key in enumerate(columns, 1):
                value = row.get(key)
                cell = result_sheet.cell(row_index, column_index, value)
                cell.font = Font(name="맑은 고딕", size=10)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if key in {"quantity", "unit_price"} and isinstance(value, (int, float)):
                    cell.number_format = "#,##0.###"

        for column_index, key in enumerate(columns, 1):
            values = [cls.COLUMN_LABELS.get(key, key)] + [str(row.get(key, "")) for row in rows]
            width = min(max(max(len(value) for value in values) + 2, 11), 32)
            result_sheet.column_dimensions[get_column_letter(column_index)].width = width
        result_sheet.row_dimensions[1].height = 24

        info_sheet = workbook.create_sheet("조회정보")
        info_sheet.sheet_view.showGridLines = False
        info_rows = [
            ("조회 모델", product_id),
            ("기준일", as_of_date or "현재 유효 기준"),
            ("조회 건수", len(rows)),
            ("생성 시각", generated_at),
            ("생성 기능", "Display BOM MCP / export_bom_excel"),
        ]
        for row_index, (label, value) in enumerate(info_rows, 1):
            label_cell = info_sheet.cell(row_index, 1, label)
            value_cell = info_sheet.cell(row_index, 2, value)
            label_cell.fill = cls.INFO_FILL
            label_cell.font = Font(name="맑은 고딕", bold=True, color="1F4E78")
            value_cell.font = Font(name="맑은 고딕")
            for cell in (label_cell, value_cell):
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center")
        info_sheet.column_dimensions["A"].width = 18
        info_sheet.column_dimensions["B"].width = 38

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
