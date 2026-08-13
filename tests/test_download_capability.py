import base64
import hashlib
import shutil
from io import BytesIO
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from mcp_server.capabilities import download


def _copy_data(tmp_path: Path) -> Path:
    source = Path(__file__).resolve().parents[1] / "data"
    target = tmp_path / "data"
    shutil.copytree(source, target, ignore=shutil.ignore_patterns("*.zip"))
    return target


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_export_bom_excel_is_valid_and_read_only(tmp_path, monkeypatch):
    data_dir = _copy_data(tmp_path)
    original = _hash(data_dir / "bom.csv")

    def get_bom(product_id, as_of_date=None):
        from services.bom_service import BomService
        frame = BomService(str(data_dir)).get_bom_explosion(product_id, as_of_date)
        return frame.to_dict(orient="records")

    monkeypatch.setattr(download, "get_bom_data", get_bom)
    result = download.export_bom_excel_data("LTA400HR01-0", "2026-08-12")
    content = base64.b64decode(result["file_data_base64"])
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert result["success"] is True
    assert result["row_count"] == workbook["BOM 조회결과"].max_row - 1
    assert workbook.sheetnames == ["BOM 조회결과", "조회정보"]
    assert workbook["조회정보"]["B3"].value == result["row_count"]
    assert _hash(data_dir / "bom.csv") == original


def test_export_design_change_report_is_valid_and_read_only(tmp_path, monkeypatch):
    data_dir = _copy_data(tmp_path)
    original = _hash(data_dir / "bom.csv")

    class Workflow:
        def generate_report(self, change_id):
            from services.ai_design_change_workflow_service import AiDesignChangeWorkflowService
            return AiDesignChangeWorkflowService(str(data_dir)).generate_report(change_id)

    monkeypatch.setattr(download, "AiDesignChangeWorkflowService", Workflow)
    result = download.export_design_change_report_data("CHG-20260810-001")
    content = base64.b64decode(result["file_data_base64"])
    document = Document(BytesIO(content))
    text = "\n".join(p.text for p in document.paragraphs)
    assert result["success"] is True
    assert "설계변경 · AI 품평 보고서" in text
    assert result["production_bom_modified"] is False
    assert _hash(data_dir / "bom.csv") == original


def test_export_bom_excel_empty_result_does_not_create_file(monkeypatch):
    monkeypatch.setattr(download, "get_bom_data", lambda *args, **kwargs: [])
    result = download.export_bom_excel_data("UNKNOWN", "2026-08-12")
    assert result == {
        "success": False,
        "message": "조회 조건에 해당하는 BOM이 없어 Excel을 생성하지 않았습니다.",
        "row_count": 0,
        "production_bom_modified": False,
    }
